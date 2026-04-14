# Python program to read from an excel file and add additional col, then write the output to excel file and close all files

from openpyxl import load_workbook, Workbook

# Open input excel file
inputWorkbook=load_workbook("Labsheet 18/input.xlsx")
inputSheet=inputWorkbook.active

# Create output excel file
outputWorkbook=Workbook()
outputSheet=outputWorkbook.active

# Copy header and add a new column
for col in range(1, inputSheet.max_column + 1):
    outputSheet.cell(row=1, column=col).value = inputSheet.cell(row=1, column=col).value
outputSheet.cell(row=1, column=inputSheet.max_column + 1).value = "Result"

# Copy data rows and calculate result
for row in range(2, inputSheet.max_row + 1):
    cgpa=inputSheet.cell(row=row, column=inputSheet.max_column).value
    for col in range(1, inputSheet.max_column + 1):
        outputSheet.cell(row=row, column=col).value = inputSheet.cell(row=row, column=col).value
        
    if cgpa >= 8:
        outputSheet.cell(row=row, column=inputSheet.max_column + 1).value = "Excellent"
    else:
        outputSheet.cell(row=row, column=inputSheet.max_column + 1).value = "Needs Improvement"

# Save the output excel file
outputWorkbook.save("Labsheet 18/output.xlsx")

# Close all files
inputWorkbook.close()
outputWorkbook.close()

print("Excel file processed successfully!")